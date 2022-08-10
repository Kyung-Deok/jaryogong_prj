from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from time import sleep
from bs4 import BeautifulSoup
from openpyxl import Workbook
import re

service = Service('../drivers/chromedriver.exe')

driver = webdriver.Chrome(service=service)

url = 'https://blcm.go.kr/stat/customizedStatic/CustomizedStaticSttst.do#'
driver.get(url)
sleep(1)

# 지역범위 설정 클릭
area_select = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[1]/div/div/span[1]/a')
area_select.click()
sleep(1)

# 서울특별시 확장 버튼 클릭
seoul_select = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[3]/div[2]/ul/li[1]/span/span[1]')
seoul_select.click()
sleep(1)

# 각 구 클릭
'''
/html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[3]/div[2]/ul/li[1]/ul/li[1]/span/span[2]

/html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[3]/div[2]/ul/li[1]/ul/li[25]/span/span[2]
'''
# for gu in range(1, 26, 1):
gu_select = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[3]/div[2]/ul/li[1]/ul/li[1]/span/span[2]')
gu_select.click()
sleep(1)

# 파일명, 컬럼명 지정
filename = 'crawling_building_usage_' + '1' + '_' + '1' + '.xlsx'
write_wb = Workbook()
write_ws = write_wb.active
write_ws['A1'] = 'purpose'
write_ws['B1'] = 'city'
write_ws['C1'] = 'gu'
write_ws['D1'] = 'dong'
write_ws['E1'] = 'total'

# 지역범위 선택 버튼 클릭
area_selection_select = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[3]/div[3]/span[1]/a')
area_selection_select.click()
sleep(1)

# 용도범위 설정버튼 클릭
purpose_select = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[1]/div/div/span[2]/a')
purpose_select.click()
sleep(1)

# 용도범위 선택
'''
1. 주거용: /html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[4]/div[2]/ul/li[1]/span/span[2]

2. 상업용 : /html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[4]/div[2]/ul/li[2]/span/span[1]
- 판매시설 : /html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[4]/div[2]/ul/li[2]/ul/li[4]/span/span[2]

3. 문교사회용 : /html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[4]/div[2]/ul/li[6]/span/span[1]
- 교육연구시설 : /html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[4]/div[2]/ul/li[6]/ul/li[4]/span/span[2]

4. 상업용 : /html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[4]/div[2]/ul/li[2]/span/span[1]
- 업무시설 : /html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[4]/div[2]/ul/li[2]/ul/li[6]/span/span[2]

5. 공공용 : /html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[4]/div[2]/ul/li[5]/span/span[1]
- 업무시설 : /html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[4]/div[2]/ul/li[5]/ul/li[1]/span/span[2]

6. 공업용 : /html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[4]/div[2]/ul/li[4]/span/span[1]
- 공장 : /html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[4]/div[2]/ul/li[4]/ul/li[1]/span/span[2]
'''
purpose_checkbox_select_1 = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[4]/div[2]/ul/li[1]/span/span[2]')
purpose_checkbox_select_1.click()
sleep(1)

# purpose_checkbox_select_2 = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[4]/div[2]/ul/li[2]/ul/li[4]/span/span[2]')
# purpose_checkbox_select_2.click()
# sleep(1)

# 용도범위 선택 버튼 클릭
purpose_selection_select = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[4]/div[3]/span[1]/a')
purpose_selection_select.click()
sleep(1)

# 조회버튼 클릭
search_select = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[2]/div[2]/div[2]/form/fieldset/div[2]/div[2]/button')
search_select.click()
sleep(1)


soup = BeautifulSoup(driver.page_source, 'html.parser')

# 가져올 행의 수
row_count = soup.select('span[id=rowCnt]')
row_count = int(str(row_count).split('>')[1].split(' ')[0])

# 용도 가져오기
'''
주거용 :
상업용 - 판매시설 :
문교사회용 - 교육연구시설 :
상업용 - 업무시설
공공용 - 업무시설
공업용 - 공장
'''
result_purpose = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[2]/div[2]/div[3]/div[2]/div[1]/table/thead/tr[1]/td[5]').text
for purpose_num in range(row_count):
    write_ws.cell(purpose_num + 2, 1).value = result_purpose

# 시 가져오기
result_city = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[2]/div[2]/div[3]/div[2]/div[1]/div/table/tbody/tr[1]/td[1]').text
for city_num in range(row_count):
    write_ws.cell(city_num + 2, 2).value = result_city


# 구 가져오기
result_gu = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[2]/div[2]/div[3]/div[2]/div[1]/div/table/tbody/tr[1]/td[2]').text
for gu_num in range(row_count):
    write_ws.cell(gu_num + 2, 3).value = result_gu

# 동 가져오기
'''
/html/body/div[2]/div[2]/div[2]/div[2]/div[3]/div[2]/div[1]/div/table/tbody/tr[1]/td[3]
/html/body/div[2]/div[2]/div[2]/div[2]/div[3]/div[2]/div[1]/div/table/tbody/tr[7]/td[3]
/html/body/div[2]/div[2]/div[2]/div[2]/div[3]/div[2]/div[1]/div/table/tbody/tr[8]/td[3]
'''
for dong_num in range(1, (row_count+1), 1):
    result_dong = driver.find_elements(By.XPATH, '/html/body/div[2]/div[2]/div[2]/div[2]/div[3]/div[2]/div[1]/div/table/tbody/tr[' + str(dong_num) + ']/td[3]')
    for dong in result_dong:
        write_ws.cell((dong_num) + 1, 4).value = dong.text

# 합계 가져오기
'''
/html/body/div[2]/div[2]/div[2]/div[2]/div[3]/div[2]/div[1]/div/table/tbody/tr[1]/td[4]
/html/body/div[2]/div[2]/div[2]/div[2]/div[3]/div[2]/div[1]/div/table/tbody/tr[2]/td[4]
'''

for total_num in range(1, (row_count+1), 1):
    result_total = driver.find_elements(By.XPATH, '/html/body/div[2]/div[2]/div[2]/div[2]/div[3]/div[2]/div[1]/div/table/tbody/tr['+ str(total_num) +']/td[4]')
    for total in result_total:
        total = total.text
        if total.find(',') == -1:
            total = int(total)
        else:
            total = total.split(',')[0] + total.split(',')[1]
            total = int(total)
        write_ws.cell((total_num) + 1, 5).value = total


write_wb.save(filename)






