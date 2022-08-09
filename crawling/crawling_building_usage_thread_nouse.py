from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from time import sleep
from bs4 import BeautifulSoup
from openpyxl import Workbook
import re
import requests
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor


service = Service('../../drivers/chromedriver.exe')
driver = webdriver.Chrome(service=service)

url = 'http://blcm.go.kr/map2/#'
driver.get(url)
sleep(1)

# 시 선택 클릭
city_select = driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[1]/div[1]/div/form[1]/ul/li[1]/select[1]')
city_select.click()
sleep(1)

# 서울시 선택
seoul_select = driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[1]/div[1]/div/form[1]/ul/li[1]/select[1]/option[2]')
seoul_select.click()
sleep(1)

# 시군구 선택 클릭
sigungu_select = driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[1]/div[1]/div/form[1]/ul/li[1]/select[2]')
sigungu_select.click()
sleep(1)

# 구 선택 - 25개의 구 선택 (강남구[2]~[25])
'''
2~25
/html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[1]/div[1]/div/form[1]/ul/li[1]/select[2]/option[2]
'''
try:
    # for gu in ['2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25']:
    for gu in ['5']:
        gu_select = driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[1]/div[1]/div/form[1]/ul/li[1]/select[2]/option['+ gu +']')
        # gu_select = driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[1]/div[1]/div/form[1]/ul/li[1]/select[2]/option[2]')
        gu_select.click()
        sleep(1)
    
        # 동 용도 클릭
        main_purpose_select = driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[1]/div[1]/div/form[1]/ul/li[1]/select[3]')
        main_purpose_select.click()
        sleep(1)
    
        # 동 용도 선택 (단독주택[2]~[30])
        '''
        단독주택: /html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[1]/div[1]/div/form[1]/ul/li[1]/select[3]/option[2]
        공동주택: /html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[1]/div[1]/div/form[1]/ul/li[1]/select[3]/option[3]
        판매시설: /html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[1]/div[1]/div/form[1]/ul/li[1]/select[3]/option[8]
        교육시설: /html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[1]/div[1]/div/form[1]/ul/li[1]/select[3]/option[11]
        업무시설: /html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[1]/div[1]/div/form[1]/ul/li[1]/select[3]/option[15]
        공장:    /html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[1]/div[1]/div/form[1]/ul/li[1]/select[3]/option[18]    
        '''
        # for pps in ['2', '3', '8', '11', '15', '18']:
        for pps in ['3']:
            purpose_select = driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[1]/div[1]/div/form[1]/ul/li[1]/select[3]/option['+ pps +']')
            # purpose_select = driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[1]/div[1]/div/form[1]/ul/li[1]/select[3]/option[2]')
            purpose_select.click()
            sleep(1)
    
            # 검색 버튼 클릭
            search_btn_select = driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[1]/div[1]/div/form[1]/ul/li[3]/span[1]/a')
            search_btn_select.click()
            sleep(5)
    
            print('check1: after search button click')
    
            # 파일명, 컬럼명 지정
            filename = 'crawling_building_usage_' + gu + '-' + pps + '.xlsx'
            write_wb = Workbook()
            write_ws = write_wb.active
            write_ws['A1'] = 'building_addr'
            write_ws['B1'] = 'purpose'
    
            print('check2: after file name setting')
    
            #
            soup = BeautifulSoup(driver.page_source, 'html.parser')
    
            print('check3: after parsing')
    
            ### try 2. 검색결과 숫자를 가져와서 10으로 나눠 페이지 수 예상하기
            total_result_num = soup.select('ul[class=resultCase] > li[class=fLeft] > span[class=txt]')
            print(total_result_num)
            t_r_num = str(total_result_num).split('>')[1].split('<')[0]
            print('check3-1: 쉼표 체크', t_r_num.find(','))
        
            # if t_r_num.find(',') in [1, 2]:
            #     t_r_num = t_r_num.split(',')[0] + t_r_num.split(',')[1]
            #     t_r_num = int(t_r_num)
            #     print('check3-2: 쉼표 없애기')
            # 
            
            # 숫자에 쉼표가 있는 경우 쉼표 제거하기 (연산 위해서)
            if t_r_num.find(',') == -1:
                t_r_num = int(t_r_num)
                print('check3-3: 쉼표 없애기 못 들어가뮤ㅠ')
            else:
                t_r_num = t_r_num.split(',')[0] + t_r_num.split(',')[1]
                t_r_num = int(t_r_num)
                print('check3-2: 쉼표 없애기')
    
            print('총 갯수', t_r_num)
            if t_r_num % 10 == 0:
                total_page_num = t_r_num//10
            else:
                total_page_num = t_r_num//10 + 1
    
            print('check4: after tpn calculate', total_page_num)
    
            # 예상한 페이지 수만큼 다음 내용 반복
            for pn in range(1, total_page_num + 1, 1):
                print('check5: after start repetition')
                print(pn)
                print(total_page_num)
    
                # 주소 가져오기
                for bnum in range(1, 11, 1):
                    bnum = str(bnum)
                    buildings = driver.find_elements(By.XPATH, '/html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[2]/div[1]/ul/li['+ bnum +']/a/span')
                    bdngs = list()
    
                    for b in buildings:
                        bdngs.append(b.text)
                        print(bdngs)
                        print('check6: after address print')
    
                    for i in range(len(bdngs)):
                        write_ws.cell(int(bnum) + 1 + ((int(pn)-1)*10), 1).value = bdngs[i]
    
                # 건물 목적 구분 가져오기
                for pnum in range(1, 11, 1):
                    pnum = str(pnum)
                    purposes = driver.find_elements(By.XPATH, '/html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[2]/div[1]/ul/li[' + pnum + ']/a')
                    pps = list()
    
                    for p in purposes:
                        pps.append(re.sub(r'\n', '*', p.text).split('*')[1].split(' ')[0])
                        print(pps)
                        print('check7: after purpose print')
    
                    for i in range(len(pps)):
                        write_ws.cell(int(pnum) + 1 + ((int(pn)-1)*10), 2).value = pps[i]
    
                # 해당 페이지에 내용이 들어있는지 확인하기
                '''
                '/html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[2]/div[1]/ul'
                '/html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[2]/div[1]/ul/li[1]'
                '''
                check_result = driver.find_elements(By.XPATH, '/html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[2]/div[1]/ul/li[1]')
                for cr in check_result:
                    print('페이지에 내용이 들어있는지 확인', cr)
    
                    if cr:
                        # 다음 페이지 버튼 누르기
                        # btn_total = soup.select('div[class=paging]')
                        # next_btn = re.sub(r'\n', '**', str(btn_total)).split('**')[-2]
                        print('check8: before next_btn click start')
                        if total_page_num >= 5:
                            next_btn = driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[2]/div[2]/a[7]')
                            next_btn.click()
                            sleep(13)
    
                        elif total_page_num == 4:
                            next_btn = driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[2]/div[2]/a[6]')
                            next_btn.click()
                            sleep(13)
    
                        elif total_page_num == 3:
                            next_btn = driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[2]/div[2]/a[5]')
                            next_btn.click()
                            sleep(13)
    
                        elif total_page_num == 2:
                            next_btn = driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div/div[1]/div/div[3]/div[2]/div[2]/a[4]')
                            next_btn.click()
                            sleep(13)
    
                        elif total_page_num == 1:
                            pass
                            sleep(13)
    
                        print('check9: after next_btn click')
                        pn = pn + 1
                        print(pn)
                        print('check10: the last part of repetition')
    
                    else:
                        pass
    
            # 엑셀 파일 저장하기
            write_wb.save(filename)

except Exception as ex:
    write_wb.save(filename)
    template = "An exception of type {0} occurred. Arguments:\n{1!r}"
    message = template.format(type(ex).__name__, ex.args)
    print(message)
    print('오류로 인한 종료', ex)


